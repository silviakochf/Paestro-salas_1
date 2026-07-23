import io
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# ── Paleta ─────────────────────────────────────────────────────
PRIMARY  = "0A192F"
HDR_SEC  = "172A45"
HDR_GRAY = "334155"
WHITE    = "FFFFFF"
GRAY1    = "F8FAFC"

# C  = Correto   → VERDE
C_BG  = "D1FAE5"; C_FG  = "065F46"
# NE = Número Errado → VERMELHO
NE_BG = "FEE2E2"; NE_FG = "991B1B"
# SN = Sem Número → AMARELO
SN_BG = "FEF9C3"; SN_FG = "854D0E"
# Pendente → CINZA
PD_BG = "F1F5F9"; PD_FG = "475569"

STATUS_MAP = {
    "C":   (C_BG,  C_FG,  "Correto"),
    "NE":  (NE_BG, NE_FG, "Número Errado"),
    "SN":  (SN_BG, SN_FG, "Sem Número"),
    None:  (PD_BG, PD_FG, "Pendente"),
}


def _fill(h): return PatternFill("solid", fgColor=h)
def _font(bold=False, color="1E293B", size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
def _border():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)
def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _header_row(ws, row, merge_to, text, bg, fg_color, size=13, row_h=30):
    ws.merge_cells(f"B{row}:{merge_to}{row}")
    c = ws[f"B{row}"]
    c.value = text; c.font = _font(True, fg_color, size)
    c.fill = _fill(bg); c.alignment = _align("center")
    ws.row_dimensions[row].height = row_h

def _col_headers(ws, row, cols, bg=HDR_GRAY, fg=WHITE, h=18):
    for ci, label in enumerate(cols, start=2):
        col = get_column_letter(ci)
        cell = ws[f"{col}{row}"]
        cell.value = label
        cell.font = _font(True, fg, 9)
        cell.fill = _fill(bg)
        cell.alignment = _align("center")
        cell.border = _border()
    ws.row_dimensions[row].height = h

def _set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def _serie_key(turma_nome: str):
    m_gt = re.match(r'^GT\s*(\d+)', turma_nome, re.IGNORECASE)
    if m_gt: return f"GT{m_gt.group(1)}"
    m_ano = re.match(r'^(\d+)', turma_nome)
    if m_ano: return m_ano.group(1)
    return turma_nome.split()[0] if turma_nome else "?"

def _serie_label(serie_key: str):
    if serie_key.startswith("GT"): return f"GT {serie_key[2:]}"
    if serie_key.isdigit(): return f"{serie_key}º ANO"
    return serie_key


def gerar_relatorio_salas_xlsx(escola: str, turmas: list, marks: dict) -> io.BytesIO:
    marks_int = {int(k): v for k, v in (marks or {}).items() if v}

    total = len(turmas)
    c_ct  = sum(1 for m in marks_int.values() if m.get("status") == "C")
    ne_ct = sum(1 for m in marks_int.values() if m.get("status") == "NE")
    sn_ct = sum(1 for m in marks_int.values() if m.get("status") == "SN")
    pend  = total - c_ct - ne_ct - sn_ct
    done  = c_ct + ne_ct + sn_ct
    pct   = round(done / total * 100) if total else 0
    now   = datetime.now().strftime("%d/%m/%Y %H:%M")

    series = sorted(
        set(_serie_key(d["turma"]) for d in turmas),
        key=lambda s: int(s[2:]) if s.startswith("GT") else int(s) if s.isdigit() else 999
    )

    wb = Workbook()

    # ═══════════════════════════════════════
    # ABA 1 — DETALHAMENTO COMPLETO
    # ═══════════════════════════════════════
    wd = wb.active
    wd.title = "Detalhamento Completo"
    wd.sheet_view.showGridLines = False
    _set_widths(wd, {"A":3,"B":5,"C":26,"D":10,"E":12,"F":14,"G":22,"H":16,"I":30,"J":3})

    _header_row(wd, 2, "I", f"CONFERÊNCIA DE SALAS — {escola.upper()}", PRIMARY, WHITE, 13, 30)
    _header_row(wd, 3, "I", f"Gerado em {now}  |  Total: {total} turmas  |  ✅ Correto: {c_ct}  |  🔴 NE: {ne_ct}  |  🟡 SN: {sn_ct}", HDR_SEC, WHITE, 9, 16)
    wd.row_dimensions[4].height = 8

    _col_headers(wd, 5, ["#","Turma","Turno","Série","Sala (sistema)","Status","Sala Real","Observação"])
    wd.freeze_panes = "B6"
    wd.auto_filter.ref = f"B5:I{5+total}"

    for ri, d in enumerate(turmas):
        row = ri + 6
        m = marks_int.get(ri, {})
        status = m.get("status")
        bg, fg, label = STATUS_MAP.get(status, STATUS_MAP[None])
        serie = _serie_label(_serie_key(d["turma"]))
        row_bg = bg if status else (GRAY1 if ri % 2 == 0 else WHITE)

        vals = [ri+1, d["turma"], d["turno"], serie, d["sala"],
                label, m.get("sala_real","") or "—", m.get("obs","") or "—"]
        for ci, v in enumerate(vals, start=2):
            col = get_column_letter(ci)
            cell = wd[f"{col}{row}"]
            cell.value = v; cell.border = _border()
            cell.alignment = _align("center" if ci in [2,4,5,6,7] else "left", wrap=(ci==9))
            if ci == 7:  # Status colorido
                cell.font = _font(True, fg, 9)
                cell.fill = _fill(bg)
            else:
                cell.font = _font(size=9)
                cell.fill = _fill(row_bg)
        wd.row_dimensions[row].height = 16

    # ═══════════════════════════════════════
    # ABA 2 — PROBLEMAS (NE e SN)
    # ═══════════════════════════════════════
    wp = wb.create_sheet("Problemas")
    wp.sheet_view.showGridLines = False
    _set_widths(wp, {"A":3,"B":5,"C":26,"D":10,"E":12,"F":14,"G":22,"H":16,"I":30,"J":3})

    _header_row(wp, 2, "I", "PROBLEMAS ENCONTRADOS — NE (Número Errado) e SN (Sem Número)", "991B1B", WHITE, 13, 30)
    _header_row(wp, 3, "I", f"{escola}  |  {now}  |  🔴 NE: {ne_ct}  |  🟡 SN: {sn_ct}", "7F1D1D", WHITE, 9, 16)
    wp.row_dimensions[4].height = 8

    cur = 5

    # Resumo por série
    wp.merge_cells(f"B{cur}:I{cur}")
    wp[f"B{cur}"].value = "RESUMO POR SÉRIE"
    wp[f"B{cur}"].font = _font(True, WHITE, 10)
    wp[f"B{cur}"].fill = _fill(HDR_GRAY)
    wp[f"B{cur}"].alignment = _align("center")
    wp.row_dimensions[cur].height = 18
    cur += 1

    _col_headers(wp, cur, ["Série","Total","✅ Correto","🔴 NE","🟡 SN","Pendente","% Problema","Situação"])
    cur += 1

    for serie in series:
        idx = [i for i, d in enumerate(turmas) if _serie_key(d["turma"]) == serie]
        t   = len(idx)
        cc  = sum(1 for i in idx if marks_int.get(i,{}).get("status") == "C")
        ne  = sum(1 for i in idx if marks_int.get(i,{}).get("status") == "NE")
        sn  = sum(1 for i in idx if marks_int.get(i,{}).get("status") == "SN")
        pp  = t - cc - ne - sn
        prob_pct = f"{round((ne + sn) / t * 100)}%" if t else "—"
        situacao = "⚠️ Atenção" if (ne + sn) > 0 else ("✅ OK" if cc == t else "⏳ Pendente")
        bg_row = NE_BG if ne > 0 else (SN_BG if sn > 0 else (C_BG if cc == t else PD_BG))
        for ci, v in enumerate([_serie_label(serie), t, cc, ne, sn, pp, prob_pct, situacao], start=2):
            cell = wp[f"{get_column_letter(ci)}{cur}"]
            cell.value = v; cell.border = _border()
            cell.font = _font(bold=(ci==2), size=9)
            cell.fill = _fill(bg_row)
            cell.alignment = _align("center" if ci != 2 else "left")
        wp.row_dimensions[cur].height = 16
        cur += 1

    # Total
    for ci, v in enumerate(["TOTAL GERAL", total, c_ct, ne_ct, sn_ct, pend, f"{round((ne_ct+sn_ct)/total*100) if total else 0}%", ""], start=2):
        cell = wp[f"{get_column_letter(ci)}{cur}"]
        cell.value = v; cell.border = _border()
        cell.font = _font(True, WHITE, 9)
        cell.fill = _fill(PRIMARY)
        cell.alignment = _align("center" if ci != 2 else "left")
    wp.row_dimensions[cur].height = 18
    cur += 2

    # Seção NE — vermelho
    ne_list = [(i, d) for i, d in enumerate(turmas) if marks_int.get(i,{}).get("status") == "NE"]
    if ne_list:
        wp.merge_cells(f"B{cur}:I{cur}")
        wp[f"B{cur}"].value = f"🔴 NE — NÚMERO ERRADO ({len(ne_list)} turma(s))"
        wp[f"B{cur}"].font = _font(True, NE_FG, 11)
        wp[f"B{cur}"].fill = _fill(NE_BG)
        wp[f"B{cur}"].alignment = _align("center")
        wp.row_dimensions[cur].height = 22; cur += 1

        _col_headers(wp, cur, ["#","Turma","Turno","Série","Sala (sistema)","Sala REAL","Observação",""], bg="991B1B")
        cur += 1

        for ri, (i, d) in enumerate(ne_list):
            m = marks_int[i]
            sala_real = m.get("sala_real","") or "⚠️ NÃO INFORMADA"
            vals = [ri+1, d["turma"], d["turno"], _serie_label(_serie_key(d["turma"])),
                    d["sala"], sala_real, m.get("obs","") or "—", ""]
            for ci, v in enumerate(vals, start=2):
                cell = wp[f"{get_column_letter(ci)}{cur}"]
                cell.value = v; cell.border = _border()
                cell.font = _font(bold=(ci==7), size=9, color=NE_FG if ci==7 else "1E293B")
                cell.fill = _fill(NE_BG)
                cell.alignment = _align("center" if ci in [2,4,5,6] else "left")
            wp.row_dimensions[cur].height = 16; cur += 1
        cur += 1

    # Seção SN — amarelo
    sn_list = [(i, d) for i, d in enumerate(turmas) if marks_int.get(i,{}).get("status") == "SN"]
    if sn_list:
        wp.merge_cells(f"B{cur}:I{cur}")
        wp[f"B{cur}"].value = f"🟡 SN — SEM NÚMERO ({len(sn_list)} turma(s))"
        wp[f"B{cur}"].font = _font(True, SN_FG, 11)
        wp[f"B{cur}"].fill = _fill(SN_BG)
        wp[f"B{cur}"].alignment = _align("center")
        wp.row_dimensions[cur].height = 22; cur += 1

        _col_headers(wp, cur, ["#","Turma","Turno","Série","Sala (sistema)","Observação","",""], bg="854D0E")
        cur += 1

        for ri, (i, d) in enumerate(sn_list):
            m = marks_int[i]
            vals = [ri+1, d["turma"], d["turno"], _serie_label(_serie_key(d["turma"])),
                    d["sala"], m.get("obs","") or "—", "", ""]
            for ci, v in enumerate(vals, start=2):
                cell = wp[f"{get_column_letter(ci)}{cur}"]
                cell.value = v; cell.border = _border()
                cell.font = _font(size=9)
                cell.fill = _fill(SN_BG)
                cell.alignment = _align("center" if ci in [2,4,5,6] else "left")
            wp.row_dimensions[cur].height = 16; cur += 1

    if not ne_list and not sn_list:
        wp.merge_cells(f"B{cur}:I{cur}")
        wp[f"B{cur}"].value = "✅ Nenhum problema! Todas as salas estão corretas."
        wp[f"B{cur}"].font = _font(italic=True, color=C_FG, size=10)
        wp[f"B{cur}"].fill = _fill(C_BG)
        wp[f"B{cur}"].alignment = _align("center")

    # ═══════════════════════════════════════
    # ABA 3 — RESUMO
    # ═══════════════════════════════════════
    ws = wb.create_sheet("Resumo")
    ws.sheet_view.showGridLines = False
    _set_widths(ws, {"A":3,"B":28,"C":16,"D":16,"E":16,"F":20,"G":3})

    _header_row(ws, 2, "F", "RESUMO — CONFERÊNCIA DE SALAS", PRIMARY, WHITE, 14, 32)
    _header_row(ws, 3, "F", f"{escola}  |  Ano Letivo {datetime.now().year}  |  Gerado em {now}", HDR_SEC, WHITE, 9, 18)
    ws.row_dimensions[4].height = 10

    cards = [
        ("B", "Total de Turmas",  total,  PRIMARY, WHITE),
        ("C", "✅ Correto",       c_ct,   C_BG,   C_FG),
        ("D", "🔴 Nº Errado",    ne_ct,  NE_BG,  NE_FG),
        ("E", "🟡 Sem Número",   sn_ct,  SN_BG,  SN_FG),
        ("F", "⏳ Pendentes",    pend,   PD_BG,  PD_FG),
    ]
    for col, label, val, bg, fg in cards:
        for r, (v, sz, it) in enumerate(
            [(label, 9, False), (val, 22, False), (f"{round(val/total*100) if total else 0}% do total", 8, True)],
            start=5
        ):
            c = ws[f"{col}{r}"]
            c.value = v; c.fill = _fill(bg)
            c.font = _font(r==6, fg, sz, it)
            c.alignment = _align("center"); c.border = _border()
        ws.row_dimensions[5].height = 18
        ws.row_dimensions[6].height = 30
        ws.row_dimensions[7].height = 16

    ws.row_dimensions[8].height = 8
    ws.merge_cells("B9:F9")
    ws["B9"].value = f"Progresso: {done} de {total} turmas verificadas ({pct}%)"
    ws["B9"].font = _font(True, WHITE, 10)
    ws["B9"].fill = _fill(PRIMARY)
    ws["B9"].alignment = _align("center")
    ws.row_dimensions[9].height = 20
    ws.row_dimensions[10].height = 8

    _header_row(ws, 11, "F", "Distribuição por Série", PRIMARY, WHITE, 11, 22)
    _col_headers(ws, 12, ["Série","Total","✅ Correto","🔴 NE","🟡 SN","Pendente"])

    for ri, serie in enumerate(series, start=13):
        idx = [i for i, d in enumerate(turmas) if _serie_key(d["turma"]) == serie]
        t   = len(idx)
        cc  = sum(1 for i in idx if marks_int.get(i,{}).get("status") == "C")
        ne  = sum(1 for i in idx if marks_int.get(i,{}).get("status") == "NE")
        sn  = sum(1 for i in idx if marks_int.get(i,{}).get("status") == "SN")
        pp  = t - cc - ne - sn
        bg = GRAY1 if ri % 2 == 0 else WHITE
        for ci, v in enumerate([_serie_label(serie), t, cc, ne, sn, pp], start=2):
            cl = ws[f"{get_column_letter(ci)}{ri}"]
            cl.value = v; cl.fill = _fill(bg)
            cl.font = _font(ci==2, size=9)
            cl.alignment = _align("center" if ci>2 else "left")
            cl.border = _border()
        ws.row_dimensions[ri].height = 16

    tr = 12 + len(series) + 1
    for ci, v in enumerate(["TOTAL", total, c_ct, ne_ct, sn_ct, pend], start=2):
        cl = ws[f"{get_column_letter(ci)}{tr}"]
        cl.value = v; cl.fill = _fill(PRIMARY)
        cl.font = _font(True, WHITE, 9)
        cl.alignment = _align("center" if ci>2 else "left")
        cl.border = _border()
    ws.row_dimensions[tr].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
