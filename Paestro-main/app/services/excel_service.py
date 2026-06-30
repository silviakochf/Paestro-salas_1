from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import io
from datetime import datetime
import pytz
import re
import unicodedata
import logging

# Configuração de Logger
logger = logging.getLogger(__name__)

def normalize_school_name(name):
    """
    Função auxiliar para normalizar nomes de escolas.
    """
    if not name:
        return ""
    nfkd_form = unicodedata.normalize('NFKD', name)
    name_ascii = "".join([c for c in nfkd_form if not unicodedata.combining(c)])
    return " ".join(name_ascii.upper().split())

def export_to_excel(classes, attendance_status, observations, html_content=None, current_user=None, periodo=None, escola_nome=None, unit_annotations=None):
    """
    Gera um arquivo Excel com a lista de presença formatada.
    Suporta alunos adicionados manualmente e anotações da unidade.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "LISTA DE PRESENÇA"

    br_tz = pytz.timezone('America/Sao_Paulo')
    current_time = datetime.now(br_tz).strftime('%d/%m/%Y %H:%M')

    # Validação de valores padrão
    safe_escola = escola_nome.upper() if escola_nome else "NÃO INFORMADO"
    safe_user = current_user.upper() if current_user else "NÃO INFORMADO"
    safe_periodo = periodo.upper() if periodo else "NÃO INFORMADO"

    header_rows = [
        ("UNIDADE:", safe_escola),
        ("RESPONSÁVEIS:", safe_user),
        ("PERÍODO:", safe_periodo),
        ("DATA E HORA:", current_time)
    ]

    # Renderiza o cabeçalho fixo
    for i, (label, value) in enumerate(header_rows, start=1):
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = Font(bold=True)
        ws[f'B{i}'] = value
        ws.merge_cells(f'B{i}:D{i}')

    current_row = len(header_rows) + 2 

    # --- Seção de Anotações da Unidade ---
    # Só renderiza se houver anotações
    if unit_annotations and len(unit_annotations) > 0:
        ws.merge_cells(f"A{current_row}:D{current_row}")
        ws[f"A{current_row}"] = "ANOTAÇÕES GERAIS DA UNIDADE:"
        ws[f"A{current_row}"].font = Font(bold=True, underline="single")
        current_row += 1
        
        for note in unit_annotations:
            ws.merge_cells(f"A{current_row}:D{current_row}")
            ws[f"A{current_row}"] = f"• {note}"
            ws[f"A{current_row}"].alignment = Alignment(wrap_text=True)
            current_row += 1
        
        current_row += 1 # Espaço extra após anotações

    # Identificar todas as turmas únicas (originais + onde houve lançamento manual)
    all_turmas = set(classes.keys())
    if attendance_status:
        all_turmas.update(attendance_status.keys())
    if observations:
        all_turmas.update(observations.keys())

    if not all_turmas:
        ws.merge_cells(f"A{current_row}:D{current_row}")
        ws[f"A{current_row}"] = "NENHUMA CHAMADA SALVA ENCONTRADA"
        ws[f"A{current_row}"].font = Font(italic=True, color="FF0000")
    else:
        # Ordena as turmas alfabeticamente
        sorted_classes = sorted(list(all_turmas))

        for turma in sorted_classes:
            # Pega alunos da lista original (importada)
            original_students = classes.get(turma, [])
            
            # Pega alunos que têm presença lançada (inclui manuais)
            attendance_students = list(attendance_status.get(turma, {}).keys())
            
            # Pega alunos que têm observação lançada (inclui manuais)
            obs_students = list(observations.get(turma, {}).keys())
            
            # Junta todos, remove duplicatas e ordena
            all_students = sorted(list(set(original_students + attendance_students + obs_students)))
            
            # Se não tiver alunos (turma vazia), pula
            if not all_students:
                continue

            # Limpa o nome da turma para exibição
            turma_display = turma.split('(')[0].strip() if '(' in turma else turma
            
            # Cabeçalho da Turma
            ws.merge_cells(f"A{current_row}:D{current_row}")
            ws[f"A{current_row}"] = f"TURMA: {turma_display.upper()}"
            ws[f"A{current_row}"].font = Font(bold=True, size=12, color="FFFFFF")
            
            # Estilo simples de fundo azul
            blue_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            ws[f"A{current_row}"].fill = blue_fill
            
            current_row += 1

            # Cabeçalhos das Colunas
            headers = ["ALUNO", "PRESENÇA", "OBSERVAÇÃO"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
            current_row += 1

            # Lista de Alunos
            for aluno in all_students:
                ws.cell(row=current_row, column=1).value = aluno
                
                # Status (P, F, FJ) - GARANTIA DE "P" SE VAZIO
                status = attendance_status.get(turma, {}).get(aluno)
                if not status: # Se for None ou string vazia
                    status = "P"
                
                ws.cell(row=current_row, column=2).value = status
                ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center')
                
                # Observações
                obs = observations.get(turma, {}).get(aluno, "")
                ws.cell(row=current_row, column=3).value = obs
                
                current_row += 1

            current_row += 2  # Espaço maior entre turmas

    # Ajuste de Largura das Colunas
    ws.column_dimensions["A"].width = 40  # Nome do Aluno
    ws.column_dimensions["B"].width = 15  # Presença
    ws.column_dimensions["C"].width = 50  # Observação

    # Ajuste de Alinhamento Global
    for row in ws.iter_rows(min_row=len(header_rows)+2):
        for cell in row:
            if not cell.alignment.horizontal: # Preserva alinhamento 'center' se já existir
                cell.alignment = Alignment(vertical='center', horizontal='left')

    # Salva em memória
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    logger.info(f"Arquivo Excel gerado para {safe_escola}")
    return output

def get_excel_filename(escola_nome=None, periodo=None, current_user=None):
    """
    Gera o nome do arquivo no formato padronizado.
    """
    def sanitize(text):
        if not text or not isinstance(text, str):
            return ""
        text = unicodedata.normalize('NFKD', text).encode('ASCII', 'ignore').decode('ASCII')
        pattern = r'[^\w\s-]'
        text = re.sub(pattern, '', text)
        text = re.sub(r'[\s\-\.]+', '_', text.strip())
        text = re.sub(r'_+', '_', text)
        return text.strip('_').upper()

    components = [
        sanitize(escola_nome) or "UNIDADE_NAO_INFORMADA",
        datetime.now().strftime('%d-%m-%Y'),
        sanitize(periodo) or "PERIODO_NAO_INFORMADO",
        sanitize(current_user) or "DUPLA_NAO_INFORMADA"
    ]
    
    filename = "_".join(filter(None, components)) + ".xlsx"
    return filename[:100]